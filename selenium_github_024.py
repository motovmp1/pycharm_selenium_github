import openpyxl
import time

path = "/home/elsys/PycharmProjects/pycharm_selenium_github/readexcel.xlsx"

# this is a object
workbook = openpyxl.load_workbook(path)

sheet = workbook.active

rows_number = sheet.max_row
cls_number = sheet.max_column

print(rows_number)
print(cls_number)
time.sleep(1)

for lines_rows in range(1, rows_number+1):
    for lines_columns in range(1, cls_number+1):
        print(sheet.cell(row=lines_rows, column=lines_columns).value, end="               ")
    print("")
